"""
Excel reading tool for Campaign Builder.

This module provides functionality for reading Excel spreadsheets (.xlsx, .xls)
using openpyxl library.
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Union

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore


@dataclass
class SheetData:
    """Data from a single Excel sheet.

    Attributes:
        name: Name of the sheet.
        headers: List of column headers.
        rows: List of row data (each row is a list of values).
        row_count: Total number of data rows in the sheet.
    """

    name: str
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)
    row_count: int = 0


@dataclass
class ExcelReadResult:
    """Result of reading an Excel file.

    Attributes:
        success: Whether the Excel file was read successfully.
        sheets: List of SheetData objects for each sheet read.
        metadata: Excel file metadata.
        error: Error message if reading failed, None otherwise.
        sheet_names: List of all sheet names in the workbook.
        column_types: Dict mapping column name to inferred type.
        truncated: Whether rows were truncated due to limits.
    """

    success: bool
    sheets: list[SheetData] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None
    sheet_names: list[str] = field(default_factory=list)
    column_types: dict[str, str] = field(default_factory=dict)
    truncated: bool = False


def _infer_column_type(values: list[Any]) -> str:
    """Infer the data type of a column based on sample values.

    Args:
        values: List of values from the column.

    Returns:
        Type string: "numeric", "text", "date", "mixed", or "empty".
    """
    # Filter out None/empty values
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]

    if not non_empty:
        return "empty"

    # Sample up to 20 values
    sample = non_empty[:20]

    numeric_count = 0
    text_count = 0
    date_count = 0

    for value in sample:
        if isinstance(value, (int, float)):
            numeric_count += 1
        elif hasattr(value, "strftime"):  # datetime-like
            date_count += 1
        elif isinstance(value, str):
            # Try to parse as number
            try:
                float(value.replace(",", ""))
                numeric_count += 1
            except (ValueError, AttributeError):
                text_count += 1
        else:
            text_count += 1

    total = len(sample)
    if date_count == total:
        return "date"
    elif numeric_count == total:
        return "numeric"
    elif text_count == total:
        return "text"
    else:
        return "mixed"


def _format_cell_value(value: Any) -> Any:
    """Format a cell value for output.

    Args:
        value: Raw cell value.

    Returns:
        Formatted value.
    """
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def _format_as_table(headers: list[str], rows: list[list[Any]], max_col_width: int = 30) -> str:
    """Format data as a markdown table.

    Args:
        headers: List of column headers.
        rows: List of row data.
        max_col_width: Maximum column width.

    Returns:
        Formatted markdown table string.
    """
    if not headers:
        return ""

    # Calculate column widths
    col_widths = []
    for i, header in enumerate(headers):
        max_width = len(str(header))
        for row in rows:
            if i < len(row):
                cell_len = len(str(row[i]))
                max_width = max(max_width, min(cell_len, max_col_width))
        col_widths.append(min(max_width, max_col_width))

    # Build header row
    header_cells = []
    for i, header in enumerate(headers):
        cell = str(header)[:col_widths[i]].ljust(col_widths[i])
        header_cells.append(cell)
    header_line = "| " + " | ".join(header_cells) + " |"

    # Build separator
    separator_cells = ["-" * w for w in col_widths]
    separator_line = "|" + "|".join(separator_cells) + "|"

    # Build data rows
    data_lines = []
    for row in rows:
        row_cells = []
        for i, width in enumerate(col_widths):
            if i < len(row):
                cell_value = str(row[i])
                if len(cell_value) > width:
                    cell_value = cell_value[: width - 3] + "..."
                cell = cell_value.ljust(width)
            else:
                cell = " " * width
            row_cells.append(cell)
        data_lines.append("| " + " | ".join(row_cells) + " |")

    return "\n".join([header_line, separator_line] + data_lines)


def read_excel(
    file_path: Union[str, Path],
    sheet_name: Optional[Union[str, int]] = None,
    max_rows: int = 1000,
) -> ExcelReadResult:
    """Read data from an Excel file.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name or 0-based index of sheet to read.
                   If None, reads the first sheet.
        max_rows: Maximum number of data rows to read (default 1000).

    Returns:
        ExcelReadResult containing sheet data and metadata.
    """
    # Check if openpyxl is available
    if openpyxl is None:
        return ExcelReadResult(
            success=False,
            error="openpyxl library is not installed. Install with: pip install openpyxl",
        )

    # Convert to Path object
    path = Path(file_path) if isinstance(file_path, str) else file_path

    # Validate file exists
    if not path.exists():
        return ExcelReadResult(
            success=False,
            error=f"File not found: {path}",
        )

    # Check file is not a directory
    if path.is_dir():
        return ExcelReadResult(
            success=False,
            error=f"Path is a directory, not a file: {path}",
        )

    # Check file extension
    suffix = path.suffix.lower()
    if suffix not in [".xlsx", ".xls", ".xlsm"]:
        return ExcelReadResult(
            success=False,
            error=f"File does not have a valid Excel extension (.xlsx, .xls, .xlsm): {path}",
        )

    # Check file is not empty
    if path.stat().st_size == 0:
        return ExcelReadResult(
            success=False,
            error=f"File is empty: {path}",
        )

    # Handle .xls files (old format)
    if suffix == ".xls":
        return ExcelReadResult(
            success=False,
            error="Old Excel format (.xls) is not supported. Please convert to .xlsx format.",
        )

    try:
        # Open workbook
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypted" in error_msg:
            return ExcelReadResult(
                success=False,
                error=f"Excel file is password protected: {path}",
            )
        return ExcelReadResult(
            success=False,
            error=f"Failed to open Excel file: {e}",
        )

    try:
        all_sheet_names = workbook.sheetnames

        # Determine which sheet(s) to read
        sheets_to_read = []
        if sheet_name is None:
            # Read first sheet
            sheets_to_read = [all_sheet_names[0]] if all_sheet_names else []
        elif isinstance(sheet_name, int):
            # Read sheet by index
            if 0 <= sheet_name < len(all_sheet_names):
                sheets_to_read = [all_sheet_names[sheet_name]]
            else:
                workbook.close()
                return ExcelReadResult(
                    success=False,
                    error=f"Sheet index {sheet_name} out of range. Available sheets: {all_sheet_names}",
                    sheet_names=all_sheet_names,
                )
        else:
            # Read sheet by name (case-insensitive)
            name_lower = sheet_name.lower()
            for sn in all_sheet_names:
                if sn.lower() == name_lower:
                    sheets_to_read = [sn]
                    break
            if not sheets_to_read:
                workbook.close()
                return ExcelReadResult(
                    success=False,
                    error=f"Sheet '{sheet_name}' not found. Available sheets: {all_sheet_names}",
                    sheet_names=all_sheet_names,
                )

        # Extract metadata
        metadata = {}
        if workbook.properties:
            props = workbook.properties
            for attr in ["title", "creator", "created", "modified", "subject"]:
                value = getattr(props, attr, None)
                if value:
                    if hasattr(value, "strftime"):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    metadata[attr] = value

        # Read sheets
        result_sheets = []
        all_column_types = {}
        truncated = False

        for sheet_name_to_read in sheets_to_read:
            sheet = workbook[sheet_name_to_read]

            # Read all rows
            all_rows = []
            for row in sheet.iter_rows():
                row_values = [_format_cell_value(cell.value) for cell in row]
                all_rows.append(row_values)

            if not all_rows:
                result_sheets.append(
                    SheetData(
                        name=sheet_name_to_read,
                        headers=[],
                        rows=[],
                        row_count=0,
                    )
                )
                continue

            # Find header row (first non-empty row)
            header_row_idx = 0
            for idx, row in enumerate(all_rows):
                if any(v for v in row if v is not None and str(v).strip()):
                    header_row_idx = idx
                    break

            # Extract headers
            raw_headers = all_rows[header_row_idx] if header_row_idx < len(all_rows) else []
            headers = []
            for i, h in enumerate(raw_headers):
                if h is None or str(h).strip() == "":
                    headers.append(get_column_letter(i + 1))
                else:
                    headers.append(str(h))

            # Extract data rows
            data_rows = all_rows[header_row_idx + 1 :]
            total_data_rows = len(data_rows)

            if len(data_rows) > max_rows:
                data_rows = data_rows[:max_rows]
                truncated = True

            # Infer column types
            for col_idx, header in enumerate(headers):
                col_values = [
                    row[col_idx] if col_idx < len(row) else None for row in data_rows
                ]
                col_type = _infer_column_type(col_values)
                all_column_types[header] = col_type

            result_sheets.append(
                SheetData(
                    name=sheet_name_to_read,
                    headers=headers,
                    rows=data_rows,
                    row_count=total_data_rows,
                )
            )

        workbook.close()

        return ExcelReadResult(
            success=True,
            sheets=result_sheets,
            metadata=metadata,
            sheet_names=all_sheet_names,
            column_types=all_column_types,
            truncated=truncated,
        )

    except Exception as e:
        workbook.close()
        return ExcelReadResult(
            success=False,
            error=f"Error reading Excel content: {e}",
        )


def format_excel_as_table(result: ExcelReadResult, sheet_index: int = 0) -> str:
    """Format Excel data as a markdown table.

    Args:
        result: ExcelReadResult from read_excel.
        sheet_index: Index of sheet to format (default 0).

    Returns:
        Formatted markdown table string.
    """
    if not result.success or not result.sheets:
        return ""

    if sheet_index >= len(result.sheets):
        return ""

    sheet = result.sheets[sheet_index]
    return _format_as_table(sheet.headers, sheet.rows)
